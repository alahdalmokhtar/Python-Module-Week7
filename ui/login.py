from PyQt6 import QtCore, QtGui, QtWidgets, uic
import os
from openpyxl import load_workbook

class new_login(object):
    def setupUi(self, new_login):
        self.window = new_login
        new_login.setObjectName("new_login")
        new_login.resize(541, 315)
        new_login.setStyleSheet("font: 700 10pt \"Times New Roman\";\n"
                                "background-color: rgb(165, 170, 169);")

        self.centralwidget = QtWidgets.QWidget(parent=new_login)
        self.centralwidget.setObjectName("centralwidget")

        self.label_3 = QtWidgets.QLabel(parent=self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(60, -5, 481, 111))
        self.label_3.setPixmap(QtGui.QPixmap('ui/logo-1.png'))
        self.label_3.setObjectName("label_3")

        self.layoutWidget = QtWidgets.QWidget(parent=self.centralwidget)
        self.layoutWidget.setGeometry(QtCore.QRect(120, 130, 216, 98))
        self.layoutWidget.setObjectName("layoutWidget")

        self.gridLayout = QtWidgets.QGridLayout(self.layoutWidget)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)

        self.label = QtWidgets.QLabel(parent=self.layoutWidget)
        self.label.setStyleSheet("font: 700 10pt \"Times New Roman\";")
        self.label.setObjectName("label")
        self.gridLayout.addWidget(self.label, 0, 0, 1, 1)

        self.lineEdit = QtWidgets.QLineEdit(parent=self.layoutWidget)
        self.lineEdit.setObjectName("lineEdit")
        self.gridLayout.addWidget(self.lineEdit, 0, 1, 1, 2)

        self.label_2 = QtWidgets.QLabel(parent=self.layoutWidget)
        self.label_2.setStyleSheet("font: 700 10pt \"Times New Roman\";")
        self.label_2.setObjectName("label_2")
        self.gridLayout.addWidget(self.label_2, 1, 0, 1, 1)

        self.lineEdit_2 = QtWidgets.QLineEdit(parent=self.layoutWidget)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.lineEdit_2.setEchoMode(QtWidgets.QLineEdit.EchoMode.Password)
        self.gridLayout.addWidget(self.lineEdit_2, 1, 1, 1, 2)

        self.errorlabel = QtWidgets.QLabel(parent=self.layoutWidget)
        self.errorlabel.setStyleSheet("color: red;")
        self.errorlabel.setObjectName("errorlabel")
        self.gridLayout.addWidget(self.errorlabel, 2, 1, 1, 2)

        self.pushButton = QtWidgets.QPushButton(parent=self.layoutWidget)
        self.pushButton.setStyleSheet("font: 700 10pt \"Times New Roman\";\n"
                                      "background-color: rgb(207, 255, 253);")
        self.pushButton.setObjectName("pushButton")
        self.gridLayout.addWidget(self.pushButton, 3, 0, 1, 2)

        self.pushButton_2 = QtWidgets.QPushButton(parent=self.layoutWidget)
        self.pushButton_2.setStyleSheet("font: 700 10pt \"Times New Roman\";\n"
                                        "background-color: rgb(207, 255, 253);")
        self.pushButton_2.setObjectName("pushButton_2")
        self.gridLayout.addWidget(self.pushButton_2, 3, 2, 1, 1)

        new_login.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(parent=new_login)
        new_login.setStatusBar(self.statusbar)

        self.retranslateUi(new_login)
        QtCore.QMetaObject.connectSlotsByName(new_login)

        # Connect buttons
        self.pushButton.clicked.connect(self.handle_login)
        self.pushButton_2.clicked.connect(QtWidgets.QApplication.quit)

    def retranslateUi(self, new_login):
        _translate = QtCore.QCoreApplication.translate
        new_login.setWindowTitle(_translate("new_login", "Login User form"))
        self.label.setText(_translate("new_login", "User name"))
        self.label_2.setText(_translate("new_login", "Password"))
        self.pushButton.setText(_translate("new_login", "Login"))
        self.pushButton_2.setText(_translate("new_login", "Exit"))

    def load_users_from_excel(self):
        try:
            wb = load_workbook(filename='Users.xlsx')
            sheet = wb.active
            users = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    username = str(row[0]).strip()
                    password = str(row[1]).strip()
                    role = str(row[2]).strip().lower() if row[2] else "user"
                    users[username] = {"password": password, "role": role}
            return users
        except FileNotFoundError:
            self.errorlabel.setText("file Users.xlsx not found")
            return {}
        except Exception as e:
            self.errorlabel.setText(f"wrong with read file {str(e)}")
            return {}
    def handle_login(self):
      self.check_login()
 
    def check_login(self):
        username = self.lineEdit.text().strip()
        password = self.lineEdit_2.text().strip()

        if not username or not password:
            self.errorlabel.setText("Please enter both username and password")
            return

        users = self.load_users_from_excel()
        if username in users:
            if users[username]["password"] == password:
                role = users[username]["role"]
                self.errorlabel.setStyleSheet("color: green;")
                self.errorlabel.setText(f"Hallo {username}! transver to admin page..")
                QtCore.QTimer.singleShot(1000, lambda: self.open_dashboard(role, username))
            else:
                self.errorlabel.setStyleSheet("color: red;")
                self.errorlabel.setText("Password is incorrect")
        else:
            self.errorlabel.setStyleSheet("color: red;")
            self.errorlabel.setText("User not found")

    def open_dashboard(self, role, username):
        if role == "admin":
            self.open_admin_form(username)
        else:
            self.open_user_form(username)

    def open_admin_form(self, username):
     try:
        """ from adminmenue import Ui_adminmenue  # Move import here
        self.pref_window = Ui_adminmenue()
        self.pref_window.show() """
        """ ui_file = os.path.join(os.path.dirname(__file__), "AdminMenue.ui")
        self.admin_window = uic.loadUi(ui_file)
        self.admin_window.setWindowTitle(f"Admin Dashboard - {username}")
        self.admin_window.show()
        QtWidgets.QApplication.instance().activeWindow().close() """
        from adminmenue import Ui_adminmenue  # Move import here
        self.admin_window = QtWidgets.QMainWindow()
        self.admin_ui = Ui_adminmenue()
        self.admin_ui.setupUi(self.admin_window)
        self.admin_window.show()
        
     except Exception as e:
        self.errorlabel.setStyleSheet("color: red;")
        self.errorlabel.setText(f"wrong the open control panel for Admin :{str(e)}")


    def open_user_form(self, username):
     try:
        """ ui_file = os.path.join(os.path.dirname(__file__), "UserMenue.ui")
        self.user_window = uic.loadUi(ui_file)
        self.user_window.setWindowTitle(f"User Dashboard - {username}")
        self.user_window.show()
        QtWidgets.QApplication.instance().activeWindow().close() """
        from usermenue import Ui_usermenue  # Move import here
        self.user_window = QtWidgets.QMainWindow()
        self.user_ui = Ui_usermenue()
        self.user_ui.setupUi(self.user_window)
        self.user_window.show()
     except Exception as e:
        self.errorlabel.setStyleSheet("color: red;")
        self.errorlabel.setText(f" wrong the open control panel for user : {str(e)}")


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    login_window = QtWidgets.QMainWindow()
    ui = new_login()
    ui.setupUi(login_window)
    login_window.show()
    sys.exit(app.exec())
