from PyQt6 import QtCore, QtGui, QtWidgets
import sys
from pathlib import Path
from openpyxl import load_workbook

# Add the parent directory to the system path to import backend modules
sys.path.append(str(Path(__file__).parent.parent))

class Ui_Interviews(object):
    def setupUi(self, Interviews):
        Interviews.setObjectName("Interviews")
        Interviews.resize(877, 540)
        Interviews.setStyleSheet("font: 10pt \"Times New Roman\";")
        self.centralwidget = QtWidgets.QWidget(parent=Interviews)
        self.centralwidget.setObjectName("centralwidget")

        # Interface layout
        self.layoutWidget = QtWidgets.QWidget(parent=self.centralwidget)
        self.layoutWidget.setGeometry(QtCore.QRect(10, 120, 150, 321))
        self.layoutWidget.setObjectName("layoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.layoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)

        self.lineEdit = QtWidgets.QLineEdit(parent=self.layoutWidget)
        self.lineEdit.setStyleSheet("color: rgb(255, 0, 0);\nbackground-color: rgb(100, 108, 102);")
        self.verticalLayout.addWidget(self.lineEdit)

        self.pushButton = QtWidgets.QPushButton(parent=self.layoutWidget)
        self.verticalLayout.addWidget(self.pushButton)

        self.pushButton_2 = QtWidgets.QPushButton(parent=self.layoutWidget)
        self.verticalLayout.addWidget(self.pushButton_2)

        self.pushButton_5 = QtWidgets.QPushButton(parent=self.layoutWidget)
        self.verticalLayout.addWidget(self.pushButton_5)

        self.bk_menue = QtWidgets.QPushButton(parent=self.layoutWidget)
        self.verticalLayout.addWidget(self.bk_menue)

        self.pushButton_4 = QtWidgets.QPushButton(parent=self.layoutWidget)
        self.verticalLayout.addWidget(self.pushButton_4)

        # Headers and labels
        self.label_2 = QtWidgets.QLabel(parent=self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(470, 20, 161, 71))
        self.label_2.setStyleSheet("font: 700 18pt \"Times New Roman\";")

        self.label_3 = QtWidgets.QLabel(parent=self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(10, 0, 151, 111))
        self.label_3.setPixmap(QtGui.QPixmap('ui/hh.png'))
        self.label_3.setScaledContents(True)

        self.label = QtWidgets.QLabel(parent=self.centralwidget)
        self.label.setGeometry(QtCore.QRect(180, -10, 281, 121))
        self.label.setStyleSheet("background-color: rgb(165, 170, 169);")
        self.label.setPixmap(QtGui.QPixmap('ui/logo-1.png'))
        self.label.setScaledContents(True)

        # Table
        self.tableWidget = QtWidgets.QTableWidget(parent=self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(180, 120, 621, 321))
        self.tableWidget.setStyleSheet("background-color: rgb(170, 170, 170);\nfont: 700 8pt \"Segoe UI\";")
        self.tableWidget.setRowCount(0)
        self.tableWidget.setColumnCount(3)
        self.tableWidget.setHorizontalHeaderLabels([
            "Name Sure Name", "Project submission", "Project arrival date"
        ])
        self.tableWidget.horizontalHeader().setMinimumSectionSize(200)

        Interviews.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(parent=Interviews)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 877, 21))
        Interviews.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(parent=Interviews)
        Interviews.setStatusBar(self.statusbar)

        self.retranslateUi(Interviews)
        QtCore.QMetaObject.connectSlotsByName(Interviews)
        
        # Initialize data storage
        self.original_data = []
        self.current_filtered_data = []
        
        # Connect buttons
        self.lineEdit.textChanged.connect(self.search_action)
        self.pushButton_2.clicked.connect(self.submitted_projects_action)
        self.pushButton_5.clicked.connect(self.projects_arrivals_action)
        self.bk_menue.clicked.connect(self.back_to_menu)
        self.pushButton_4.clicked.connect(self.exit_app)
        
        # Load data automatically
        self.load_interview_data()

    def retranslateUi(self, Interviews):
        _translate = QtCore.QCoreApplication.translate
        Interviews.setWindowTitle(_translate("Interviews", "Interviews"))
        self.pushButton.setText(_translate("Interviews", "Search"))
        self.pushButton_2.setText(_translate("Interviews", "Submitted Projects"))
        self.pushButton_5.setText(_translate("Interviews", "Projects Arrivals"))
        self.bk_menue.setText(_translate("Interviews", "Back Menue"))
        self.pushButton_4.setText(_translate("Interviews", "Exit"))
        self.label_2.setText(_translate("Interviews", "Interview Page"))

    def load_interview_data(self):
        """Load data from Excel file and populate the table, ignoring empty/None values"""
        try:
            file_path = "Interviews.xlsx"
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            
            self.original_data = []
            
            # Read headers first
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            if len(headers) >= 3:  # Ensure we have all required columns
                self.original_data.append(headers)
            
            # Read data rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Convert row to list and check for valid data
                row_data = [
                    str(cell) if cell is not None else ""
                    for cell in row[:3]  # Only take first 3 columns
                ]
                
                # Only add row if it has valid name and at least one other field
                if row_data[0].strip() and (row_data[1].strip() or row_data[2].strip()):
                    self.original_data.append(row_data)
            
            # Display all data initially
            self.current_filtered_data = self.original_data.copy()
            self.display_data(self.current_filtered_data)
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(None, "Error", f"Failed to load interview data:\n{str(e)}")

    def display_data(self, data):
        """Display data in the table widget, ensuring only valid data is shown"""
        self.tableWidget.setRowCount(0)
        
        if not data or len(data) < 2:  # Need at least headers + one row
            return
            
        # Skip header if it exists
        start_row = 1 if len(data) > 1 else 0
        
        self.tableWidget.setRowCount(len(data) - start_row)
        
        for row_idx, row_data in enumerate(data[start_row:]):
            for col_idx in range(3):  # Only show first 3 columns
                item = QtWidgets.QTableWidgetItem(str(row_data[col_idx]) if col_idx < len(row_data) else QtWidgets.QTableWidgetItem(""))
                self.tableWidget.setItem(row_idx, col_idx, item)

    def search_action(self):
        """Filter table based on search text"""
        search_text = self.lineEdit.text().strip().lower()
        
        if not search_text:
            self.display_data(self.original_data)
            return
            
        filtered = [self.original_data[0]]  # Keep headers
        
        for row in self.original_data[1:]:
            if any(search_text in str(cell).lower() for cell in row[:3]):
                filtered.append(row)
        
        self.display_data(filtered)

    def submitted_projects_action(self):
        """Filter to show only rows with valid submitted projects"""
        try:
            if not self.original_data:
                QtWidgets.QMessageBox.warning(None, "Warning", "No data loaded")
                return
                
            filtered = [self.original_data[0]]  # Keep headers
            
            for row in self.original_data[1:]:
                if len(row) > 1 and row[1].strip() and row[1].strip().lower() not in ['no', 'none', 'nan']:
                    filtered.append(row)
            
            self.display_data(filtered)
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(None, "Error", f"Failed to filter submitted projects:\n{str(e)}")

    def projects_arrivals_action(self):
        """Filter to show only rows with valid arrived projects"""
        try:
            if not self.original_data:
                QtWidgets.QMessageBox.warning(None, "Warning", "No data loaded")
                return
                
            filtered = [self.original_data[0]]  # Keep headers
            
            for row in self.original_data[1:]:
                if len(row) > 2 and row[2].strip() and row[2].strip().lower() not in ['no', 'none', 'nan']:
                    filtered.append(row)
            
            self.display_data(filtered)
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(None, "Error", f"Failed to filter arrived projects:\n{str(e)}")

    def back_to_menu(self):
        from preferencesadmin import PreferencesAdminWindow
        self.admin_window = PreferencesAdminWindow()
        self.admin_window.show()

    def exit_app(self):
        QtWidgets.QApplication.quit()


def mentor():
    app = QtWidgets.QApplication(sys.argv)
    interview_window = QtWidgets.QMainWindow()
    ui = Ui_Interviews()
    ui.setupUi(interview_window)
    interview_window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    mentor()