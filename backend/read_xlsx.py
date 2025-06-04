# This script reads an Excel file and prints its contents.
# 
from openpyxl import load_workbook

def read_xlsx(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        if all(cell is None for cell in row):  # skip empty row
            continue

        data.append(row)
    return data
    print(data)
if __name__ == "__main__":
    #read_xlsx("Applications.xlsx")  # Replace with your actual file path
    #read_xlsx("Users.xlsx")
    #read_xlsx("Interview.xlsx")
    read_xlsx("Mentors.xlsx")